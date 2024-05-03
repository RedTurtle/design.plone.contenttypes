# -*- coding: utf-8 -*-
from DateTime import DateTime
from design.plone.contenttypes.utils import text_in_block
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from plone import api
from Products.Five import BrowserView

import io


FLAG = '<i class="fa-solid fa-check"></i>'


class CheckServizi(BrowserView):
    cds = None

    def is_anonymous(self):
        return api.user.is_anonymous()

    def get_canale_accesso_info(self, servizio):
        canale_fisico = getattr(servizio, "canale_fisico", None)
        canale_digitale = text_in_block(getattr(servizio, "canale_digitale", None))
        canale_digitale_link = getattr(servizio, "canale_digitale_link", None)
        if canale_digitale and canale_digitale_link and canale_fisico:
            return "D e F"
        elif canale_digitale and canale_digitale_link:
            return "D"
        elif canale_fisico:
            return "F"
        # elif canale_digitale or canale_digitale_link:
        #     return "&frac12;D"
        return None

    def information_dict(self, servizio):
        return {
            "title": getattr(servizio, "title"),
            "description": getattr(servizio, "description", None),
            "condizioni_di_servizio": getattr(servizio, "condizioni_di_servizio", None),
            "tassonomia_argomenti": getattr(servizio, "tassonomia_argomenti", None),
            "a_chi_si_rivolge": text_in_block(
                getattr(servizio, "a_chi_si_rivolge", None)
            ),
            "come_si_fa": text_in_block(getattr(servizio, "come_si_fa", None)),
            "cosa_si_ottiene": text_in_block(
                getattr(servizio, "cosa_si_ottiene", None)
            ),
            "canale_accesso": self.get_canale_accesso_info(servizio),
            "cosa_serve": text_in_block(getattr(servizio, "cosa_serve", None)),
            "tempi_e_scadenze": text_in_block(
                getattr(servizio, "tempi_e_scadenze", None)
            ),
            "ufficio_responsabile": getattr(servizio, "ufficio_responsabile", None),
            "contact_info": getattr(servizio, "contact_info", None),
        }

    def plone2volto(self, url):
        portal_url = api.portal.get().absolute_url()
        frontend_domain = api.portal.get_registry_record(
            "volto.frontend_domain", default=""
        )
        if frontend_domain and url.startswith(portal_url):
            return url.replace(portal_url, frontend_domain, 1)
        return url

    def get_relation_title(self, information_dict, label):
        result = ""
        for item in information_dict[label]:
            if not item:
                continue
            if not item.to_object:
                continue
            result = result + " " + item.to_object.title
        return result

    def get_servizi(self, full_report=False):
        if self.is_anonymous():
            return []
        pc = api.portal.get_tool("portal_catalog")

        # show_inactive ha sempre avuto una gestione... particolare! aggiungo ai
        # kw effectiveRange = DateTime() che è quello che fa Products.CMFPlone
        # nel CatalogTool.py
        query = {
            "portal_type": "Servizio",
            "review_state": "published",
            # "show_inactive": False,
        }
        brains = pc(query, **{"effectiveRange": DateTime()})
        results = {}
        for brain in brains:
            servizio = brain.getObject()
            # if safe_hasattr(servizio, "condizioni_di_servizio") and getattr(
            #     servizio, "condizioni_di_servizio"
            # ):
            #     continue

            # Se chiediamo di vedere anche le condizioni di servizio, lo teniamo
            self.cds = self.request.get("condizioni_di_servizio", None)
            information_dict = self.information_dict(servizio)
            if not self.cds:
                del information_dict["condizioni_di_servizio"]

            if all(information_dict.values()):
                continue

            parent = servizio.aq_inner.aq_parent
            if parent.title not in results:
                results[parent.title] = {
                    "url": self.plone2volto(parent.absolute_url()),
                    "children": [],
                }

            tassonomia_argomenti = self.get_relation_title(
                information_dict, "tassonomia_argomenti"
            )
            ufficio_responsabile = self.get_relation_title(
                information_dict, "ufficio_responsabile"
            )

            contatti = self.get_relation_title(information_dict, "contact_info")

            if full_report:
                results[parent.title]["children"].append(
                    {
                        "title": servizio.title,
                        "url": self.plone2volto(servizio.absolute_url()),
                        "data": {
                            "title": information_dict.get("title") or "",
                            "description": information_dict.get("description") or "",
                            "condizioni_di_servizio": information_dict.get(
                                "condizioni_di_servizio"
                            )
                            or "",
                            "tassonomia_argomenti": tassonomia_argomenti or "",
                            "a_chi_si_rivolge": information_dict.get("a_chi_si_rivolge")
                            or "",
                            "come_si_fa": information_dict.get("come_si_fa") or "",
                            "cosa_si_ottiene": information_dict.get("cosa_si_ottiene")
                            or "",
                            "canale_accesso": information_dict.get("canale_accesso")
                            or "",
                            "cosa_serve": information_dict.get("cosa_serve") or "",
                            "tempi_e_scadenze": information_dict.get("tempi_e_scadenze")
                            or "",
                            "ufficio_responsabile": ufficio_responsabile or "",
                            "contact_info": contatti or "",
                        },
                    }
                )
            else:
                results[parent.title]["children"].append(
                    {
                        "title": servizio.title,
                        "url": self.plone2volto(servizio.absolute_url()),
                        "data": {
                            "title": information_dict.get("title") and FLAG or "",
                            "description": information_dict.get("description")
                            and FLAG
                            or "",
                            "condizioni_di_servizio": information_dict.get(
                                "condizioni_di_servizio"
                            )
                            and FLAG
                            or "",
                            "tassonomia_argomenti": tassonomia_argomenti and FLAG or "",
                            "a_chi_si_rivolge": information_dict.get("a_chi_si_rivolge")
                            and FLAG
                            or "",
                            "come_si_fa": information_dict.get("come_si_fa")
                            and FLAG
                            or "",
                            "cosa_si_ottiene": information_dict.get("cosa_si_ottiene")
                            and FLAG
                            or "",
                            "canale_accesso": information_dict.get("canale_accesso")
                            or "",
                            "cosa_serve": information_dict.get("cosa_serve")
                            and FLAG
                            or "",
                            "tempi_e_scadenze": information_dict.get("tempi_e_scadenze")
                            and FLAG
                            or "",
                            "ufficio_responsabile": ufficio_responsabile and FLAG or "",
                            "contact_info": contatti and FLAG or "",
                        },
                    }
                )

        results = dict(sorted(results.items()))
        for key in results:
            results[key]["children"].sort(key=lambda x: x["title"])
        return results


class DownloadCheckServizi(CheckServizi):
    CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def __call__(self):
        HEADER = [
            "Titolo",
            "Descrizione",
            "Argomenti",
            "A chi è rivolto",
            "Come fare per",
            "Cosa si ottiene",
            "Canale di accesso",
            "Cosa serve",
            "Tempi e scadenze",
            "Unità org. responsabile",
            "Contatti",
        ]

        cds = self.request.form.get("condizioni_di_servizio", None)
        if cds:
            EMPTY_ROW = [""] * 12
            HEADER.insert(2, "Condizioni di servizio")
        else:
            EMPTY_ROW = [""] * 11

        full_report = self.request.form.get("full", False)
        servizi = self.get_servizi(full_report)

        data = []
        for category in servizi:
            data.append([category] + [""] * 10 + [servizi[category]["url"]])
            data.append(HEADER)

            if full_report:
                for servizio in servizi[category]["children"]:
                    dati_servizio = [
                        servizio["title"],
                        servizio["data"]["description"] or "",
                        servizio["data"]["tassonomia_argomenti"] or "",
                        servizio["data"]["a_chi_si_rivolge"] or "",
                        servizio["data"]["come_si_fa"] or "",
                        servizio["data"]["cosa_si_ottiene"] or "",
                        servizio["data"]["canale_accesso"] or "",
                        servizio["data"]["cosa_serve"] or "",
                        servizio["data"]["tempi_e_scadenze"] or "",
                        servizio["data"]["ufficio_responsabile"] or "",
                        servizio["data"]["contact_info"] or "",
                        servizio["url"],
                    ]
                    if cds:
                        condizioni_di_servizio = (
                            servizio["data"]["condizioni_di_servizio"] or ""
                        )
                        dati_servizio.insert(2, condizioni_di_servizio)
                    data.append(dati_servizio)
            else:
                for servizio in servizi[category]["children"]:
                    dati_servizio = [
                        servizio["title"],
                        servizio["data"]["description"] and "V" or "",
                        servizio["data"]["tassonomia_argomenti"] and "V" or "",
                        servizio["data"]["a_chi_si_rivolge"] and "V" or "",
                        servizio["data"]["come_si_fa"] and "V" or "",
                        servizio["data"]["cosa_si_ottiene"] and "V" or "",
                        servizio["data"]["canale_accesso"] and "V" or "",
                        servizio["data"]["cosa_serve"] and "V" or "",
                        servizio["data"]["tempi_e_scadenze"] and "V" or "",
                        servizio["data"]["ufficio_responsabile"] and "V" or "",
                        servizio["data"]["contact_info"] and "V" or "",
                        servizio["url"],
                    ]
                    if cds:
                        condizioni_di_servizio = (
                            servizio["data"]["condizioni_di_servizio"] or ""
                        )
                        dati_servizio.insert(2, condizioni_di_servizio)
                    data.append(dati_servizio)
            data.append(EMPTY_ROW)
            data.append(EMPTY_ROW)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Servizi"
        link_font = Font(underline="single", color="0563C1")
        link_fill = PatternFill(fill_type="solid", fgColor="DDDDDD")
        alignment = Alignment(horizontal="center", vertical="top")

        for i, row in enumerate(data, start=1):
            have_url = row[0] != "Titolo" and bool(row[0])
            if have_url:
                url = row.pop()

            sheet.append(row)

            if row[0] == "Titolo":
                for index, cell in enumerate(sheet[i]):
                    cell.fill = link_fill
                    if index != 0:
                        column_letter = get_column_letter(cell.column)
                        sheet.column_dimensions[column_letter].width = 20

            if have_url:
                for index, cell in enumerate(sheet[i]):
                    if index == 0:
                        cell = sheet.cell(row=i, column=1)
                        cell.font = link_font
                        cell.alignment = cell.alignment.copy(wrapText=True)
                        cell.hyperlink = url
                        column_letter = get_column_letter(cell.column)
                        sheet.column_dimensions[column_letter].width = 40
                    else:
                        cell.alignment = alignment

        bytes_io = io.BytesIO()
        workbook.save(bytes_io)
        data = bytes_io.getvalue()
        self.request.response.setHeader("Content-Length", len(data))
        self.request.RESPONSE.setHeader("Content-Type", self.CT)
        self.request.response.setHeader(
            "Content-Disposition",
            "attachment; filename=check_servizi.xlsx",
        )
        return data
