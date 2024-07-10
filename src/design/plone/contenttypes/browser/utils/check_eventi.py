from DateTime import DateTime
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from plone import api
from Products.Five import BrowserView

import io


FLAG = '<i class="fa-solid fa-check"></i>'


class CheckEventi(BrowserView):
    cds = None

    def is_anonymous(self):
        return api.user.is_anonymous()

    def information_dict(self, evento):
        prezzo = getattr(evento, "prezzo", "")
        res = [x.get("text", "") for x in prezzo["blocks"].values()]
        if not [x for x in res if x]:
            prezzo = ""

        luoghi_correlati = False
        if getattr(evento, "luoghi_correlati", None):
            luoghi_correlati = True
        elif getattr(evento, "geolocation", None):
            if getattr(evento.geolocation, "latitude", "") and getattr(
                evento.geolocation, "longitude", ""
            ):
                luoghi_correlati = True

        return {
            "description": getattr(evento, "description", "").strip(),
            "effective_date": getattr(evento, "effective_date", None),
            "luoghi_correlati": luoghi_correlati,
            "prezzo": prezzo,
            "contact_info": getattr(evento, "contact_info", None),
        }

    def plone2volto(self, url):
        portal_url = api.portal.get().absolute_url()
        frontend_domain = api.portal.get_registry_record(
            "volto.frontend_domain", default=""
        )
        if frontend_domain and url.startswith(portal_url):
            return url.replace(portal_url, frontend_domain, 1)
        return url

    def get_eventi(self):
        if self.is_anonymous():
            return []
        pc = api.portal.get_tool("portal_catalog")

        query = {
            "portal_type": "Event",
            "review_state": "published",
            "effectiveRange": DateTime(),
        }
        brains = pc(query)
        results = {}
        for brain in brains:
            evento = brain.getObject()

            information_dict = self.information_dict(evento)
            if all(information_dict.values()):
                continue

            parent = evento.aq_inner.aq_parent
            if parent.title not in results:
                results[parent.title] = {
                    "url": self.plone2volto(parent.absolute_url()),
                    "children": [],
                }
            results[parent.title]["children"].append(
                {
                    "title": evento.title,
                    "description": information_dict.get("description") and FLAG or "",
                    "url": self.plone2volto(evento.absolute_url()),
                    "data": {
                        "effective_date": information_dict.get("effective_date")
                        and FLAG
                        or "",
                        "luoghi_correlati": information_dict.get("luoghi_correlati")
                        and FLAG
                        or "",
                        "prezzo": information_dict.get("prezzo") and FLAG or "",
                        "contact_info": information_dict.get("contact_info")
                        and FLAG
                        or "",
                    },
                }
            )

        results = dict(sorted(results.items()))
        for key in results:
            results[key]["children"].sort(key=lambda x: x["title"])

        return results


class DownloadCheckEventi(CheckEventi):
    CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def __call__(self):
        HEADER = [
            "Titolo",
            "Descrizione",
            "Data",
            "Luogo",
            "Costo",
            "Contatti",
        ]

        EMPTY_ROW = [""] * 3

        eventi = self.get_eventi()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Check Persone"
        header_font = Font(bold=True)
        section_link_font = Font(underline="single", color="0563C1", size=14)
        link_fill = PatternFill(fill_type="solid", fgColor="DDDDDD")
        link_font = Font(underline="single", color="0563C1")
        section_fill = PatternFill(fill_type="solid", fgColor="E9E9E9")
        alignment = Alignment(horizontal="center", vertical="top", wrapText=True)

        section_row_height = int(14 * 1.5)

        for category, category_data in eventi.items():
            section_url = category_data["url"]
            section_title = category
            section_row = [section_title, "", ""]
            sheet.append(section_row)
            section_cell = sheet.cell(row=sheet.max_row, column=1)

            section_cell.alignment = alignment
            section_cell.hyperlink = section_url
            sheet.merge_cells(
                start_row=sheet.max_row,
                start_column=1,
                end_row=sheet.max_row,
                end_column=3,
            )  # noqa
            for row in sheet.iter_rows(
                min_row=sheet.max_row, max_row=sheet.max_row, min_col=1, max_col=3
            ):  # noqa
                sheet.row_dimensions[row[0].row].height = section_row_height
                for cell in row:
                    cell.fill = section_fill
                    cell.font = section_link_font

            sheet.append(HEADER)
            for col in range(1, len(HEADER) + 1):
                header_cell = sheet.cell(row=sheet.max_row, column=col)
                header_cell.fill = link_fill
                header_cell.font = header_font

            for col in sheet.columns:
                column_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[column_letter].width = 35

            for evento in category_data["children"]:
                title_url = evento["url"]
                dati_evento = [
                    evento["title"],
                    "X" if evento["description"] else "",
                    "X" if evento["data"]["effective_date"] else "",
                    "X" if evento["data"]["luoghi_correlati"] else "",
                    "X" if evento["data"]["prezzo"] else "",
                    "X" if evento["data"]["contact_info"] else "",
                ]
                row = dati_evento
                sheet.append(row)

                title_cell = sheet.cell(row=sheet.max_row, column=1)
                check_cell = sheet.cell(row=sheet.max_row, column=2)
                check_cell.alignment = check_cell.alignment.copy(horizontal="center")
                title_cell.hyperlink = title_url
                title_cell.font = link_font
                column_letter_unit = get_column_letter(title_cell.column)
                sheet.column_dimensions[column_letter_unit].width = 60

            sheet.append(EMPTY_ROW)
            sheet.append(EMPTY_ROW)

        bytes_io = io.BytesIO()
        workbook.save(bytes_io)
        data = bytes_io.getvalue()
        self.request.response.setHeader("Content-Length", len(data))
        self.request.RESPONSE.setHeader("Content-Type", self.CT)
        self.request.response.setHeader(
            "Content-Disposition",
            "attachment; filename=check_eventi.xlsx",
        )
        return data
