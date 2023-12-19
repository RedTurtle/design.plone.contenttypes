from DateTime import DateTime
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from plone import api
from Products.Five import BrowserView
from z3c.relationfield import RelationValue
from zope.component import getUtility
from zope.intid.interfaces import IIntIds

import io
import transaction


FLAG = '<i class="fa-solid fa-check"></i>'


class CheckNotizie(BrowserView):
    cds = None

    def is_anonymous(self):
        return api.user.is_anonymous()

    def information_dict(self, notizia):
        descrizione_estesa = getattr(notizia, "descrizione_estesa", "")
        res = [x.get("text", "") for x in descrizione_estesa["blocks"].values()]
        if not [x for x in res if x]:
            descrizione_estesa = ""

        return {
            "descrizione_estesa": descrizione_estesa,
            "effective_date": getattr(notizia, "effective_date", None),
            "a_cura_di": getattr(notizia, "a_cura_di", None),
        }

    def plone2volto(self, url):
        portal_url = api.portal.get().absolute_url()
        frontend_domain = api.portal.get_registry_record(
            "volto.frontend_domain", default=""
        )
        if frontend_domain and url.startswith(portal_url):
            return url.replace(portal_url, frontend_domain, 1)
        return url

    def get_notizie(self):
        if self.is_anonymous():
            return []
        pc = api.portal.get_tool("portal_catalog")

        query = {
            "portal_type": "News Item",
            "review_state": "published",
            "effectiveRange": DateTime(),
        }

        brains = pc(query)
        results = {}
        for brain in brains:
            notizia = brain.getObject()

            information_dict = self.information_dict(notizia)

            if all(information_dict.values()):
                continue

            parent = notizia.aq_inner.aq_parent
            if parent.title not in results:
                results[parent.title] = {
                    "url": self.plone2volto(parent.absolute_url()),
                    "path": "/".join(parent.getPhysicalPath()),
                    "children": [],
                }

            results[parent.title]["children"].append(
                {
                    "title": notizia.title,
                    "descrizione_estesa": information_dict.get("descrizione_estesa")
                    and FLAG
                    or "",
                    "url": self.plone2volto(notizia.absolute_url()),
                    "UID": notizia.UID(),
                    "data": {
                        "effective_date": information_dict.get("effective_date")
                        and FLAG
                        or "",
                        "a_cura_di": information_dict.get("a_cura_di") and FLAG or "",
                    },
                }
            )

        results = dict(sorted(results.items()))
        for key in results:
            results[key]["children"].sort(key=lambda x: x["title"])

        return results


class SetACuraDi(CheckNotizie):
    def __call__(self):
        # news_folder = self.request.get("path")
        came_from = self.request.get("came_from")
        uids = self.request.get("uids")
        news_folder = self.request.get("path")
        portal_id = api.portal.get().id
        path_ufficio = f"/{portal_id}" + self.request.get("path_ufficio")
        ufficio = api.content.get(path=path_ufficio)

        if not ufficio:
            self.context.plone_utils.addPortalMessage(
                f"Impossibile trovare l'ufficio {self.request.get('path_ufficio')}",
                "error",
            )

            self.request.RESPONSE.redirect(came_from)
            return

        # same query above, plus news UIDS; use UIDS 'cause trying to get brains
        # by path give me strange results
        query = {
            "portal_type": "News Item",
            "review_state": "published",
            "effectiveRange": DateTime(),
            "UID": uids,
        }
        pc = api.portal.get_tool("portal_catalog")
        brains = pc(query)
        intids = getUtility(IIntIds)
        ufficio = api.content.get(path=path_ufficio)

        for idx, brain in enumerate(brains):
            notizia = brain.getObject()

            information_dict = self.information_dict(notizia)
            if all(information_dict.values()):
                continue

            if not getattr(notizia, "a_cura_di", None):
                notizia.a_cura_di = [RelationValue(intids.getId(ufficio))]
                notizia.reindexObject()

            if idx % 10 == 0:
                transaction.commit()

        office_path = "/".join(ufficio.getPhysicalPath())
        self.context.plone_utils.addPortalMessage(
            f'Impostato l\'ufficio "{ufficio.title}" ({office_path}) nelle news della cartella {news_folder}',  # noqa
            "info",
        )
        self.request.RESPONSE.redirect(came_from)
        return


class DownloadCheckNotizie(CheckNotizie):
    CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def __call__(self):
        HEADER = [
            "Titolo",
            "Descrizione estesa",
            "Data di pubblicazione",
            "A cura di",
        ]

        EMPTY_ROW = [""] * 3

        notizie = self.get_notizie()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Check Persone"
        header_font = Font(bold=True)
        section_link_font = Font(underline="single", color="0563C1", size=14)
        link_fill = PatternFill(fill_type="solid", fgColor="DDDDDD")
        link_font = Font(underline="single", color="0563C1")
        section_fill = PatternFill(fill_type="solid", fgColor="E9E9E9")
        alignment = Alignment(horizontal="center", vertical="top", wrapText=True)

        section_row_height = int(14 * 1.5)

        for category, category_data in notizie.items():
            section_url = category_data["url"]
            section_title = category
            section_row = [section_title, "", ""]
            sheet.append(section_row)
            section_cell = sheet.cell(row=sheet.max_row, column=1)

            section_cell.alignment = alignment
            section_cell.hyperlink = section_url
            sheet.merge_cells(
                start_row=sheet.max_row,
                start_column=1,
                end_row=sheet.max_row,
                end_column=3,
            )  # noqa
            for row in sheet.iter_rows(
                min_row=sheet.max_row, max_row=sheet.max_row, min_col=1, max_col=3
            ):  # noqa
                sheet.row_dimensions[row[0].row].height = section_row_height
                for cell in row:
                    cell.fill = section_fill
                    cell.font = section_link_font

            sheet.append(HEADER)
            for col in range(1, len(HEADER) + 1):
                header_cell = sheet.cell(row=sheet.max_row, column=col)
                header_cell.fill = link_fill
                header_cell.font = header_font

            for col in sheet.columns:
                column_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[column_letter].width = 35

            for notizia in category_data["children"]:
                title_url = notizia["url"]
                dati_notizia = [
                    notizia["title"],
                    "X" if notizia["descrizione_estesa"] else "",
                    "X" if notizia["data"]["effective_date"] else "",
                    "X" if notizia["data"]["a_cura_di"] else "",
                ]
                row = dati_notizia
                sheet.append(row)

                title_cell = sheet.cell(row=sheet.max_row, column=1)
                check_cell = sheet.cell(row=sheet.max_row, column=2)
                check_cell.alignment = check_cell.alignment.copy(horizontal="center")
                title_cell.hyperlink = title_url
                title_cell.font = link_font
                column_letter_unit = get_column_letter(title_cell.column)
                sheet.column_dimensions[column_letter_unit].width = 60

            sheet.append(EMPTY_ROW)
            sheet.append(EMPTY_ROW)

        bytes_io = io.BytesIO()
        workbook.save(bytes_io)
        data = bytes_io.getvalue()
        self.request.response.setHeader("Content-Length", len(data))
        self.request.RESPONSE.setHeader("Content-Type", self.CT)
        self.request.response.setHeader(
            "Content-Disposition",
            "attachment; filename=check_notizie.xlsx",
        )
        return data
