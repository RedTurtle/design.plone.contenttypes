from DateTime import DateTime
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from plone import api
from Products.Five import BrowserView
from zope.globalrequest import getRequest
from zope.component import getMultiAdapter
from plone.restapi.interfaces import ISerializeToJsonSummary

import io


FLAG = '<i class="fa-solid fa-check"></i>'


class CheckUnitaOrganizzative(BrowserView):
    cds = None

    def is_anonymous(self):
        return api.user.is_anonymous()

    def get_relations(self, obj, field):
        return api.relation.get(source=obj, relationship=field, unrestricted=False)

    def get_related_objects(self, obj, field):
        """ """
        items = []
        relations = self.get_relations(obj, field)

        for rel in relations:
            rel_obj = rel.to_object
            if rel_obj is not None:
                summary = getMultiAdapter(
                    (rel_obj, getRequest()), ISerializeToJsonSummary
                )()
                items.append(summary)
        return sorted(items, key=lambda k: k["title"])

    def information_dict(self, uo):
        sede_ref = None
        sede_ref = self.get_related_objects(uo, "sede")
        if sede_ref:
            sede_ref = sede_ref[0]

        competenze = getattr(uo, "competenze", "")
        if competenze:
            res = [x.get("text", "") for x in competenze["blocks"].values()]
            if not [x for x in res if x]:
                competenze = ""
        return {
            "description": getattr(uo, "description", "").strip(),
            "competenze": competenze,
            "sede": sede_ref,
            "contact_info": getattr(uo, "contact_info", None),
        }

    def plone2volto(self, url):
        portal_url = api.portal.get().absolute_url()
        frontend_domain = api.portal.get_registry_record(
            "volto.frontend_domain", default=""
        )
        if frontend_domain and url.startswith(portal_url):
            return url.replace(portal_url, frontend_domain, 1)
        return url

    def get_uos(self):
        if self.is_anonymous():
            return []
        pc = api.portal.get_tool("portal_catalog")

        # show_inactive ha sempre avuto una gestione... particolare! aggiungo ai
        # kw effectiveRange = DateTime() che è quello che fa Products.CMFPlone
        # nel CatalogTool.py
        query = {
            "portal_type": "UnitaOrganizzativa",
            "review_state": "published",
        }
        brains = pc(query, **{"effectiveRange": DateTime()})
        results = {}
        for brain in brains:
            uo = brain.getObject()

            information_dict = self.information_dict(uo)

            if all(information_dict.values()):
                continue

            parent = uo.aq_inner.aq_parent
            if parent.title not in results:
                results[parent.title] = {
                    "url": self.plone2volto(parent.absolute_url()),
                    "children": [],
                }
            results[parent.title]["children"].append(
                {
                    "title": uo.title,
                    "description": information_dict.get("description") and FLAG or "",
                    "url": self.plone2volto(uo.absolute_url()),
                    "data": {
                        "competenze": information_dict.get("competenze") and FLAG or "",
                        "sede": information_dict.get("sede") and FLAG or "",
                        "contact_info": information_dict.get("contact_info")
                        and FLAG
                        or "",
                    },
                }
            )

        results = dict(sorted(results.items()))
        for key in results:
            results[key]["children"].sort(key=lambda x: x["title"])

        return results


class DownloadCheckUnitaOrganizzative(CheckUnitaOrganizzative):
    CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def __call__(self):
        HEADER = ["Titolo", "Descrizione", "Competenze", "Sede", "Contatti"]

        EMPTY_ROW = [""] * 3

        unita_organizzative = self.get_uos()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Check Persone"
        header_font = Font(bold=True)
        section_link_font = Font(underline="single", color="0563C1", size=14)
        link_fill = PatternFill(fill_type="solid", fgColor="DDDDDD")
        link_font = Font(underline="single", color="0563C1")
        section_fill = PatternFill(fill_type="solid", fgColor="E9E9E9")
        alignment = Alignment(horizontal="center", vertical="top", wrapText=True)

        section_row_height = int(14 * 1.5)

        for category, category_data in unita_organizzative.items():
            section_url = category_data["url"]
            section_title = category
            section_row = [section_title, "", ""]
            sheet.append(section_row)
            section_cell = sheet.cell(row=sheet.max_row, column=1)

            section_cell.alignment = alignment
            section_cell.hyperlink = section_url
            sheet.merge_cells(
                start_row=sheet.max_row,
                start_column=1,
                end_row=sheet.max_row,
                end_column=3,
            )  # noqa
            for row in sheet.iter_rows(
                min_row=sheet.max_row, max_row=sheet.max_row, min_col=1, max_col=3
            ):  # noqa
                sheet.row_dimensions[row[0].row].height = section_row_height
                for cell in row:
                    cell.fill = section_fill
                    cell.font = section_link_font

            sheet.append(HEADER)
            for col in range(1, len(HEADER) + 1):
                header_cell = sheet.cell(row=sheet.max_row, column=col)
                header_cell.fill = link_fill
                header_cell.font = header_font

            for col in sheet.columns:
                column_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[column_letter].width = 35

            for uo in category_data["children"]:
                title_url = uo["url"]
                dati_uo = [
                    uo["title"],
                    "X" if uo["description"] else "",
                    "X" if uo["data"]["competenze"] else "",
                    "X" if uo["data"]["sede"] else "",
                    "X" if uo["data"]["contact_info"] else "",
                ]
                row = dati_uo
                sheet.append(row)

                title_cell = sheet.cell(row=sheet.max_row, column=1)
                check_cell = sheet.cell(row=sheet.max_row, column=2)
                check_cell.alignment = check_cell.alignment.copy(horizontal="center")
                title_cell.hyperlink = title_url
                title_cell.font = link_font
                column_letter_unit = get_column_letter(title_cell.column)
                sheet.column_dimensions[column_letter_unit].width = 60

            sheet.append(EMPTY_ROW)
            sheet.append(EMPTY_ROW)

        bytes_io = io.BytesIO()
        workbook.save(bytes_io)
        data = bytes_io.getvalue()
        self.request.response.setHeader("Content-Length", len(data))
        self.request.RESPONSE.setHeader("Content-Type", self.CT)
        self.request.response.setHeader(
            "Content-Disposition",
            "attachment; filename=check_unita_organizzative.xlsx",
        )
        return data
