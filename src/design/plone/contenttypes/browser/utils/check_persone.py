from DateTime import DateTime
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from plone import api
from Products.Five import BrowserView
from zope.globalrequest import getRequest
from zope.component import getMultiAdapter
from plone.restapi.interfaces import ISerializeToJsonSummary

import io

from Acquisition import aq_inner
from zope.component import getUtility
from zope.intid.interfaces import IIntIds
from zope.security import checkPermission
from zc.relation.interfaces import ICatalog


FLAG = '<i class="fa-solid fa-check"></i>'


class CheckPersone(BrowserView):
    cds = None

    def is_anonymous(self):
        return api.user.is_anonymous()

    def get_relations(self, obj, field):
        return api.relation.get(source=obj, relationship=field, unrestricted=False)

    def get_related_objects(self, obj, field):
        """ """
        items = []
        relations = self.get_relations(obj, field)

        for rel in relations:
            rel_obj = rel.to_object
            if rel_obj is not None:
                summary = getMultiAdapter(
                    (rel_obj, getRequest()), ISerializeToJsonSummary
                )()
                items.append(summary)
        return sorted(items, key=lambda k: k["title"])

    def back_references(self, source_object, attribute_name):
        catalog = getUtility(ICatalog)
        intids = getUtility(IIntIds)
        result = []

        for rel in catalog.findRelations(
            dict(
                to_id=intids.getId(aq_inner(source_object)),
                from_attribute=attribute_name,
            )
        ):
            obj = intids.queryObject(rel.from_id)
            if obj is not None and checkPermission("zope2.View", obj):
                result.append(obj)
        return result

    def information_dict(self, persona):
        uo_refs = self.back_references(persona, "responsabile")
        uo_refs.extend(self.back_references(persona, "persone_struttura"))

        incarichi_persona = ""
        if persona.incarichi_persona:
            relations = self.get_related_objects(persona, "incarichi_persona")
            if relations:
                rel_data = relations[0]

                if (
                    rel_data["data_inizio_incarico"]
                    and rel_data["title"].strip()
                    and rel_data["tipologia_incarico"]
                ):
                    incarichi_persona = FLAG

        return {
            "incarichi_persona": incarichi_persona,
            "contact_info": getattr(persona, "contact_info", None),
            "organizzazione_riferimento": uo_refs,
        }

    def plone2volto(self, url):
        portal_url = api.portal.get().absolute_url()
        frontend_domain = api.portal.get_registry_record(
            "volto.frontend_domain", default=""
        )
        if frontend_domain and url.startswith(portal_url):
            return url.replace(portal_url, frontend_domain, 1)
        return url

    def get_persone(self):
        if self.is_anonymous():
            return []
        pc = api.portal.get_tool("portal_catalog")

        # show_inactive ha sempre avuto una gestione... particolare! aggiungo ai
        # kw effectiveRange = DateTime() che è quello che fa Products.CMFPlone
        # nel CatalogTool.py
        query = {
            "portal_type": "Persona",
            "review_state": "published",
        }
        brains = pc(query, **{"effectiveRange": DateTime()})
        results = {}
        for brain in brains:
            persona = brain.getObject()

            information_dict = self.information_dict(persona)

            if all(information_dict.values()):
                continue

            parent = persona.aq_inner.aq_parent
            if parent.title not in results:
                results[parent.title] = {
                    "url": self.plone2volto(parent.absolute_url()),
                    "children": [],
                }

            results[parent.title]["children"].append(
                {
                    "title": persona.title,
                    "url": self.plone2volto(persona.absolute_url()),
                    "data": {
                        "organizzazione_riferimento": information_dict.get(
                            "organizzazione_riferimento"
                        )
                        and FLAG
                        or "",
                        "incarichi_persona": information_dict.get("incarichi_persona"),
                        "contact_info": information_dict.get("contact_info")
                        and FLAG
                        or "",
                    },
                }
            )

        results = dict(sorted(results.items()))
        for key in results:
            results[key]["children"].sort(key=lambda x: x["title"])

        return results


class DownloadCheckPersone(CheckPersone):
    CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def __call__(self):
        HEADER = ["Titolo", "Incarichi", "Unità org. di riferimento", "Contatti"]

        EMPTY_ROW = [""] * 3

        persone = self.get_persone()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Check Persone"
        header_font = Font(bold=True)
        section_link_font = Font(underline="single", color="0563C1", size=14)
        link_fill = PatternFill(fill_type="solid", fgColor="DDDDDD")
        link_font = Font(underline="single", color="0563C1")
        section_fill = PatternFill(fill_type="solid", fgColor="E9E9E9")
        alignment = Alignment(horizontal="center", vertical="top", wrapText=True)

        section_row_height = int(14 * 1.5)

        for category, category_data in persone.items():
            section_url = category_data["url"]
            section_title = category
            section_row = [section_title, "", ""]
            sheet.append(section_row)
            section_cell = sheet.cell(row=sheet.max_row, column=1)

            section_cell.alignment = alignment
            section_cell.hyperlink = section_url
            sheet.merge_cells(
                start_row=sheet.max_row,
                start_column=1,
                end_row=sheet.max_row,
                end_column=3,
            )  # noqa
            for row in sheet.iter_rows(
                min_row=sheet.max_row, max_row=sheet.max_row, min_col=1, max_col=3
            ):  # noqa
                sheet.row_dimensions[row[0].row].height = section_row_height
                for cell in row:
                    cell.fill = section_fill
                    cell.font = section_link_font

            sheet.append(HEADER)
            for col in range(1, len(HEADER) + 1):
                header_cell = sheet.cell(row=sheet.max_row, column=col)
                header_cell.fill = link_fill
                header_cell.font = header_font

            for col in sheet.columns:
                column_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[column_letter].width = 35

            for persona in category_data["children"]:
                title_url = persona["url"]
                dati_persona = [
                    persona["title"],
                    "X" if persona["data"]["incarichi_persona"] else "",
                    "X" if persona["data"]["organizzazione_riferimento"] else "",
                    "X" if persona["data"]["contact_info"] else "",
                ]
                row = dati_persona
                sheet.append(row)

                title_cell = sheet.cell(row=sheet.max_row, column=1)
                check_cell = sheet.cell(row=sheet.max_row, column=2)
                check_cell.alignment = check_cell.alignment.copy(horizontal="center")
                title_cell.hyperlink = title_url
                title_cell.font = link_font
                column_letter_unit = get_column_letter(title_cell.column)
                sheet.column_dimensions[column_letter_unit].width = 60

            sheet.append(EMPTY_ROW)
            sheet.append(EMPTY_ROW)

        bytes_io = io.BytesIO()
        workbook.save(bytes_io)
        data = bytes_io.getvalue()
        self.request.response.setHeader("Content-Length", len(data))
        self.request.RESPONSE.setHeader("Content-Type", self.CT)
        self.request.response.setHeader(
            "Content-Disposition",
            "attachment; filename=check_persone.xlsx",
        )
        return data
