from openpyxl import Workbook
from openpyxl.styles import Font
from plone import api
from plone.namedfile.browser import Download
from Products.Five import BrowserView
from zope.interface import implementer
from zope.interface import Interface
from zope.publisher.interfaces import IPublishTraverse
from zope.publisher.interfaces import NotFound

import io


class IUsersSummaryView(Interface):
    """marker interface"""


@implementer(IUsersSummaryView)
class UsersSummaryView(BrowserView):
    """"""

    def __call__(self):
        return

    def absolute_url(self):
        """
        Needed for plone.namedfile >= 6.4.0 with canonical header
        """
        return f"{self.context.absolute_url()}/{self.__name__}"


@implementer(IPublishTraverse)
class UsersSummaryDownload(Download):
    def __init__(self, context, request):
        super().__init__(context=context, request=request)
        self.data = None

    def publishTraverse(self, request, name):
        super().publishTraverse(request=request, name=name)

        if name == "users.xlsx":
            # XXX: al momento l'estrazione avviene a partire dai gruppi,
            #      questo perchè il sito potrebbe essere connesso ad un LDAP
            #      aziendale contenente molti utenti e non ci interessa estrarli
            #      tutti

            users = {}
            groups = [
                g.getId()
                for g in api.group.get_groups()
                if g.getId() not in ["AuthenticatedUsers"]
            ]
            for group in api.group.get_groups():
                for userid in group.getGroupMemberIds():
                    if userid not in users:
                        user = api.user.get(userid=userid)
                        if user:
                            users[userid] = {
                                "username": user.getUserName(),
                                "email": user.getProperty("email"),
                                "fullname": user.getProperty("fullname"),
                            }
                            for g in user.getGroups():
                                users[userid][g] = "x"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = f"Redazione {api.portal.get().Title()}"
            header_font = Font(bold=True)
            sheet.append(
                [
                    "USERNAME",
                    "EMAIL",
                    "NOME",
                ]
            )
            sheet.cell(row=1, column=1).font = header_font
            sheet.cell(row=1, column=2).font = header_font
            sheet.cell(row=1, column=3).font = header_font
            for idx, g in enumerate(groups):
                sheet.cell(row=1, column=4 + idx).value = g
                sheet.cell(row=1, column=4 + idx).font = header_font

            for row, user in enumerate(
                sorted(users.values(), key=lambda u: u["username"])
            ):
                sheet.append(
                    [
                        user["username"],
                        user["email"],
                        user["fullname"],
                    ]
                )
                for col, g in enumerate(groups):
                    sheet.cell(row=row + 2, column=4 + col).value = user.get(g)

            bytes_io = io.BytesIO()
            workbook.save(bytes_io)
            self.data = bytes_io.getvalue()
            self.request.response.setHeader("Content-Length", len(self.data))
            self.request.RESPONSE.setHeader(
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.request.response.setHeader(
                "Content-Disposition",
                "attachment; filename=users.xlsx",
            )
        else:
            self.data = None

        return self

    # def _getFile(self):
    def __call__(self):
        if not self.data:
            raise NotFound(self, "b", self.request)
        return self.data
